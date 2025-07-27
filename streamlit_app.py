import streamlit as st
import pandas as pd
import requests
from io import StringIO
import os
from dotenv import load_dotenv
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
import shutil
import time
from datetime import datetime, timedelta
import base64
import json
import io

# Configuration
load_dotenv()
BASE_URL = "https://emis.dhis2nigeria.org.ng/dhis/api"
HEADERS = {
    "Content-Type": "application/json",
    "Authorization": f"ApiToken {os.getenv('DHIS2_API_TOKEN', 'd2pat_DAULytLK7GUBSlfknZ6ffVxIfifC8T9c3885349470')}"
}
DATASET_UIDS = [
    "MLTLNUmvS8r", "uSw8GwPO417", "W36yBpVEUkH",
    "pJydop5Fpsz", "XERITHzkeSI", "RlfDdEEZ317"
]
PERIOD = "LAST_YEAR"
OUTPUT_FOLDER = "./data"
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

MAPPING_FILE = os.path.join(OUTPUT_FOLDER, "dataset_uid_mapping.csv")
COMBINED_FILE = os.path.join(OUTPUT_FOLDER, "dataset_completion_report_EMIS.csv")
USER_FILE = os.path.join(OUTPUT_FOLDER, "dhis2_users.csv")
REPORT_EXCEL = os.path.join(OUTPUT_FOLDER, "State_school_user_upload_status_report.xlsx")
NO_UPLOAD_FILE = os.path.join(OUTPUT_FOLDER, "logged_in_no_upload.xlsx")
STATUS_CHART_PATH = os.path.join(OUTPUT_FOLDER, "LGA_Status_Chart.png")

# State options
STATE_OPTIONS = [
    {"uid": "LEVEL-st3hrLkzuMb;FHlOerryBjk", "name": "Abia State"},
    {"uid": "LEVEL-st3hrLkzuMb;OgjFloqKoqk", "name": "Adamawa State"},
    {"uid": "LEVEL-st3hrLkzuMb;qLiKWoddwFu", "name": "Akwa-Ibom State"},
    {"uid": "LEVEL-st3hrLkzuMb;Nko8QFDmYmq", "name": "Anambra state"},
    {"uid": "LEVEL-st3hrLkzuMb;ziJ3yxfgb3m", "name": "Bauchi State"},
    {"uid": "LEVEL-st3hrLkzuMb;MXrZyuS9E7A", "name": "Benue State"},
    {"uid": "LEVEL-st3hrLkzuMb;RLySnRCE1Gy", "name": "Borno State"},
    {"uid": "LEVEL-st3hrLkzuMb;ns3vF75Y0bF", "name": "Bayelsa State"},
    {"uid": "LEVEL-st3hrLkzuMb;caG44DzHu6F", "name": "Cross River State"},
    {"uid": "LEVEL-st3hrLkzuMb;m0rZG06GdPe", "name": "Delta State"},
    {"uid": "LEVEL-st3hrLkzuMb;xWSEoKmrbBW", "name": "Ebonyi State"},
    {"uid": "LEVEL-st3hrLkzuMb;aMQcvAoEFh0", "name": "Edo State"},
    {"uid": "LEVEL-st3hrLkzuMb;iilma7EajGc", "name": "Ekiti State"},
    {"uid": "LEVEL-st3hrLkzuMb;Quac4RHRtaZ", "name": "Enugu State"},
    {"uid": "LEVEL-st3hrLkzuMb;HYCMnXqLDPV", "name": "Federal Capital Territory"},
    {"uid": "LEVEL-st3hrLkzuMb;bSfaEpPFa9Y", "name": "Gombe State"},
    {"uid": "LEVEL-st3hrLkzuMb;FmOhtDnhdwU", "name": "Imo State"},
    {"uid": "LEVEL-st3hrLkzuMb;MJVVi73YayJ", "name": "Jigawa State"},
    {"uid": "LEVEL-st3hrLkzuMb;tjLatcokcel", "name": "Kaduna State"},
    {"uid": "LEVEL-st3hrLkzuMb;M689V9w3Gs3", "name": "Kebbi State"},
    {"uid": "LEVEL-st3hrLkzuMb;cTIw3RXOLCQ", "name": "Kano State"},
    {"uid": "LEVEL-st3hrLkzuMb;S7Vs7ifJKlh", "name": "Kogi State"},
    {"uid": "LEVEL-st3hrLkzuMb;uKlacgs9ykR", "name": "Katsina State"},
    {"uid": "LEVEL-st3hrLkzuMb;jReUW6NCPkL", "name": "Kwara State"},
    {"uid": "LEVEL-st3hrLkzuMb;H2ZhSMudlMI", "name": "Lagos State"},
    {"uid": "LEVEL-st3hrLkzuMb;gzLOszDWdqM", "name": "Nasarawa State"},
    {"uid": "LEVEL-st3hrLkzuMb;RYEnw3sMDyE", "name": "Niger State"},
    {"uid": "LEVEL-st3hrLkzuMb;fBInDsbaQHO", "name": "Ogun State"},
    {"uid": "LEVEL-st3hrLkzuMb;r3IK5qdHsZ6", "name": "Ondo State"},
    {"uid": "LEVEL-st3hrLkzuMb;hfNPq5F4mjr", "name": "Osun State"},
    {"uid": "LEVEL-st3hrLkzuMb;yx3QJHm86vW", "name": "Oyo State"},
    {"uid": "LEVEL-st3hrLkzuMb;TFY8aaVkCtV", "name": "Plateau State"},
    {"uid": "LEVEL-st3hrLkzuMb;BmWTbiMgEai", "name": "Rivers State"},
    {"uid": "LEVEL-st3hrLkzuMb;Gq37IyyjUfj", "name": "Sokoto State"},
    {"uid": "LEVEL-st3hrLkzuMb;jXngIDniC8t", "name": "Taraba State"},
    {"uid": "LEVEL-st3hrLkzuMb;Ym1fEhWFWYI", "name": "Yobe State"},
    {"uid": "LEVEL-st3hrLkzuMb;FmH6buccgqx", "name": "Zamfara State"}
]

# Initialize session state
if 'report_df' not in st.session_state:
    st.session_state.report_df = None
if 'violations_df' not in st.session_state:
    st.session_state.violations_df = None
if 'chart_generated' not in st.session_state:
    st.session_state.chart_generated = False
if 'users_df' not in st.session_state:
    st.session_state.users_df = None
if 'raw_report_data' not in st.session_state:
    st.session_state.raw_report_data = None
if 'auto_refresh' not in st.session_state:
    st.session_state.auto_refresh = False
if 'refresh_interval' not in st.session_state:
    st.session_state.refresh_interval = 300
if 'notifications' not in st.session_state:
    st.session_state.notifications = []
if 'advanced_filters' not in st.session_state:
    st.session_state.advanced_filters = {
        'status': [],
        'dataset': [],
        'user_count_min': '',
        'user_count_max': '',
        'last_login_start': '',
        'last_login_end': ''
    }
if 'export_options' not in st.session_state:
    st.session_state.export_options = {
        'format': 'excel',
        'include_violations': True,
        'selected_columns': []
    }

# Load UID-to-Name Mapping
try:
    uid_map_df = pd.read_csv(MAPPING_FILE, dtype=str)
    if not {'dataset_uids', 'dataset_names'}.issubset(uid_map_df.columns):
        raise ValueError("Mapping file must contain 'dataset_uids' and 'dataset_names' columns.")
except FileNotFoundError:
    uid_map_df = pd.DataFrame({'dataset_uids': DATASET_UIDS, 'dataset_names': ['Dataset ' + uid for uid in DATASET_UIDS]})
    uid_map_df.to_csv(MAPPING_FILE, index=False)

# Helper Functions
def fetch_dataset_report(uid, org_units):
    params = {
        "dimension": [f"dx:{uid}.ACTUAL_REPORTS;{uid}.EXPECTED_REPORTS", f"ou:{org_units}"],
        "filter": f"pe:{PERIOD}",
        "tableLayout": "true",
        "columns": "dx",
        "rows": "ou",
        "skipRounding": "false",
        "completedOnly": "false",
        "hideEmptyRows": "true",
        "showHierarchy": "true"
    }
    try:
        response = requests.get(f"{BASE_URL}/analytics.csv", headers=HEADERS, params=params, timeout=120)
        response.raise_for_status()
        df = pd.read_csv(StringIO(response.text))
        df["datasetuid"] = uid
        dataset_name = uid_map_df.loc[uid_map_df['dataset_uids'] == uid, 'dataset_names'].values[0]
        df["datasets_name"] = dataset_name
        return df
    except requests.RequestException as e:
        st.error(f"Error fetching dataset {uid}: {str(e)}")
        return None

def fetch_users():
    try:
        if os.path.exists(USER_FILE):
            return pd.read_csv(USER_FILE)
    except Exception as e:
        st.warning(f"Couldn't load user data from file: {e}")

    url = f"{BASE_URL}/users.json?fields=id,name,username,userGroups[name],userRoles[name],lastLogin,organisationUnits[ancestors[name],name,id,level]&paging=true&pageSize=10000"
    seen_ids = set()
    users = []
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    while url:
        status_text.text(f"Fetching users from {url.split('?')[0]}...")
        response = requests.get(url, headers=HEADERS, timeout=60)
        if response.status_code != 200:
            st.error("Failed to fetch users")
            return None
        
        data = response.json()
        total_users = data.get("pager", {}).get("total", 1)
        fetched_users = len(users)
        
        for user in data.get("users", []):
            uid = user.get("id")
            if not uid or uid in seen_ids:
                continue
            seen_ids.add(uid)
            
            user_groups = ", ".join([g["name"] for g in user.get("userGroups", [])])
            user_roles = ", ".join([r["name"] for r in user.get("userRoles", [])])
            school_paths = []
            school_uids = []
            
            for ou in user.get("organisationUnits", []):
                if ou.get("level") == 5:  # Filter for level 5 (schools)
                    path = " > ".join([a["name"] for a in ou.get("ancestors", [])] + [ou["name"]])
                    school_paths.append(path)
                    school_uids.append(ou.get("id", ""))
            
            if school_uids:  # Only include users with level 5 OUs
                users.append({
                    "id": uid,
                    "name": user.get("name", ""),
                    "username": user.get("username", ""),
                    "userGroups": user_groups,
                    "userRoles": user_roles,
                    "lastLogin": user.get("lastLogin", ""),
                    "schoolPath": ", ".join(school_paths),
                    "schoolUID": ", ".join(school_uids)
                })
        
        progress = min(1.0, len(users) / total_users)
        progress_bar.progress(progress)
        url = data.get("pager", {}).get("nextPage")
    
    progress_bar.empty()
    status_text.empty()
    
    if users:
        users_df = pd.DataFrame(users)
        users_df.to_csv(USER_FILE, index=False)
        return users_df
    return None

def get_orgunit_name(orgunit_id):
    url = f"{BASE_URL}/33/organisationUnits/{orgunit_id}.json?fields=name"
    try:
        res = requests.get(url, headers=HEADERS, timeout=10)
        res.raise_for_status()
        return res.json().get("name", "Unknown")
    except:
        return "Unknown"

def get_validation_violations_batch(batch, report_df):
    results = {}
    for orgunit_id, dataset_uids in batch:
        orgunit_name = get_orgunit_name(orgunit_id)
        st.write(f"Checking: {orgunit_name} ({orgunit_id}) - Datasets: {dataset_uids}")
        
        violations = []
        matching_row = report_df[report_df["School ID"] == orgunit_id].iloc[0] if not report_df.empty and orgunit_id in report_df["School ID"].values else None
        state = matching_row["State"] if matching_row is not None else "Unknown"
        lga = matching_row["LGA"] if matching_row is not None else "Unknown"
        ward = matching_row["Ward"] if matching_row is not None else "Unknown"
        dataset_name = matching_row["datasets_name"] if matching_row is not None else "Unknown"
        
        for dataset_uid in dataset_uids:
            url = f"{BASE_URL}/33/validation/dataSet/{dataset_uid}.json?pe=2024&ou={orgunit_id}"
            try:
                res = requests.get(url, headers=HEADERS, timeout=15)
                res.raise_for_status()
                violations.extend(res.json().get("validationRuleViolations", []))
            except requests.exceptions.RequestException as e:
                st.warning(f"Failed for {orgunit_id} with {dataset_uid}: {e}")
        
        if not violations:
            st.write("No violations found.")
        else:
            st.write(f"{len(violations)} violation(s) found")
            for v in violations:
                rule = v.get("validationRule", {})
                rule_name = rule.get("name", "Unnamed Rule")
                st.write(f"- {rule_name}")
        
        results[(orgunit_id, tuple(dataset_uids))] = {
            "violations": violations,
            "state": state,
            "lga": lga,
            "ward": ward,
            "dataset_name": dataset_name
        }
    return results

def generate_report(selected_state):
    st.write("Fetching dataset reports...")
    all_dfs = []
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for i, uid in enumerate(DATASET_UIDS):
        status_text.text(f"Fetching dataset {i+1}/{len(DATASET_UIDS)}...")
        df = fetch_dataset_report(uid, selected_state)
        if df is not None and not df.empty:
            all_dfs.append(df)
        progress_bar.progress((i + 1) / len(DATASET_UIDS))
    
    progress_bar.empty()
    status_text.empty()
    
    if not all_dfs:
        st.error("No datasets fetched")
        return None
    
    final_df = pd.concat(all_dfs, ignore_index=True)
    final_df.to_csv(COMBINED_FILE, index=False)
    
    # Process the data
    final_df["organisationunitid_original"] = final_df["organisationunitid"]
    final_df["organisationunitid"] = final_df["organisationunitid"].str.strip().str.lower().str.replace(r'[^a-z0-9]', '', regex=True)
    
    actual_cols = [col for col in final_df.columns if "Actual reports" in col]
    final_df["Total_Actual"] = final_df[actual_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1)
    
    return final_df

def generate_full_report():
    # Try to load users from file first
    st.session_state.users_df = fetch_users()
    
    if st.session_state.users_df is None:
        st.error("No user data available. Please fetch users first.")
        return None
    
    if st.session_state.raw_report_data is None:
        if os.path.exists(COMBINED_FILE):
            st.session_state.raw_report_data = pd.read_csv(COMBINED_FILE)
        else:
            st.error("No dataset reports available. Please fetch datasets first.")
            return None
    
    with st.spinner("Generating full report..."):
        report_df = st.session_state.raw_report_data.copy()
        users_df = st.session_state.users_df.copy()
        
        # Process user data
        users_df["schoolUID_original"] = users_df["schoolUID"]
        users_df["schoolUID"] = users_df["schoolUID"].str.strip().str.lower().str.replace(r'[^a-z0-9]', '', regex=True)
        
        uploaded_schools = set(report_df[report_df["Total_Actual"] > 0]["organisationunitid"])
        all_schools = set(report_df["organisationunitid"])
        
        report_rows = []
        for school_id in sorted(all_schools):
            school_data = report_df[report_df["organisationunitid"] == school_id].iloc[0]
            school_name = school_data["organisationunitname"]
            original_school_id = school_data["organisationunitid_original"]
            dataset_uid = school_data.get("datasetuid", "")
            dataset_name = school_data.get("datasets_name", "")
            
            school_users = users_df[users_df["schoolUID"] == school_id].copy()
            school_users["schoolUID"] = school_users["schoolUID_original"]
            
            if school_users.empty:
                state = school_data.get("orgunitlevel2", "")
                lga = school_data.get("orgunitlevel3", "")
                ward = school_data.get("orgunitlevel4", "")
                status = "🚫 No user account"
                last_login = ""
            else:
                first_user = school_users.iloc[0]
                state = school_data.get("orgunitlevel2", "")
                lga = school_data.get("orgunitlevel3", "")
                ward = school_data.get("orgunitlevel4", "")
                has_login = school_users["lastLogin"].str.strip().ne("").any()
                has_upload = school_id in uploaded_schools
                if has_upload:
                    status = "✅ Logged in and uploaded data"
                elif has_login:
                    status = "⚠️ Logged in, no data upload"
                else:
                    status = "❌ User exists, yet to login"
                if len(school_users) > 1:
                    status += f" (👥 {len(school_users)} users)"
                last_logins = pd.to_datetime(school_users["lastLogin"], errors='coerce', utc=True).dropna()
                last_login = last_logins.max().strftime('%Y-%m-%d %H:%M:%S UTC') if not last_logins.empty else ""
            
            report_rows.append({
                "State": state,
                "LGA": lga,
                "Ward": ward,
                "School ID": original_school_id,
                "School Name": school_name,
                "datasetuid": dataset_uid,
                "datasets_name": dataset_name,
                "Status": status,
                "User Count": len(school_users),
                "Usernames": ", ".join(school_users["username"]),
                "Last Login": last_login
            })
        
        report_df = pd.DataFrame(report_rows)
        ordered_cols = ["State", "LGA", "Ward", "School ID", "School Name", "datasetuid", "datasets_name", "Status", "User Count", "Usernames", "Last Login"]
        report_df = report_df[ordered_cols]
        
        # Save Excel file
        temp_excel = os.path.join(OUTPUT_FOLDER, "temp_report.xlsx")
        try:
            report_df.to_excel(temp_excel, index=False, sheet_name="Report")
            wb = load_workbook(temp_excel)
            ws_report = wb["Report"]
            
            # Format Excel
            for cell in ws_report[1]:
                cell.font = Font(bold=True)
            for column_cells in ws_report.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                ws_report.column_dimensions[column_cells[0].column_letter].width = max_length + 2
            
            status_colors = {
                "✅ Logged in and uploaded data": "C6EFCE",
                "⚠️ Logged in, no data upload": "FFF2CC",
                "❌ User exists, yet to login": "F8CBAD",
                "🚫 No user account": "D9D9D9"
            }
            status_col = [cell.value for cell in ws_report[1]].index("Status") + 1
            for row in ws_report.iter_rows(min_row=2, max_row=ws_report.max_row):
                status = row[status_col - 1].value
                if isinstance(status, str):
                    for key, color in status_colors.items():
                        if status.startswith(key):
                            fill = PatternFill(start_color=color, fill_type="solid")
                            for cell in row:
                                cell.fill = fill
                            break
            
            wb.save(temp_excel)
            if os.path.exists(REPORT_EXCEL):
                os.remove(REPORT_EXCEL)
            shutil.move(temp_excel, REPORT_EXCEL)
        except Exception as e:
            st.error(f"Failed to save Excel file: {e}")
            if os.path.exists(temp_excel):
                os.remove(temp_excel)
            return None
        
        return report_df

def generate_chart(report_df):
    report_df["Clean Status"] = report_df["Status"].apply(
        lambda s: "Logged in & uploaded" if s.startswith("✅") else
                  "Logged in, no upload" if s.startswith("⚠️") else
                  "User exists, no login" if s.startswith("❌") else
                  "No user account"
    )
    status_counts = report_df.groupby(["LGA", "Clean Status"]).size().unstack(fill_value=0)
    sorted_lgas = status_counts.sum(axis=1).sort_values(ascending=False)
    top_lgas = sorted_lgas.head(20).index if len(sorted_lgas) > 20 else sorted_lgas.index
    status_counts = status_counts.loc[top_lgas]
    
    fig, ax = plt.subplots(figsize=(16, 9))
    status_counts.plot(kind='bar', stacked=True, colormap='Set2', ax=ax)
    num_lgas = len(status_counts)
    title = f"All LGAs by School Upload & Login Status ({num_lgas} total)" if num_lgas < 20 else "Top 20 LGAs by School Upload & Login Status"
    plt.title(title, fontsize=16)
    plt.xlabel("LGA", fontsize=12)
    plt.ylabel("Number of Schools", fontsize=12)
    plt.xticks(rotation=30, ha='right')
    plt.legend(title="Status", bbox_to_anchor=(1.02, 1), loc='upper left')
    plt.tight_layout()
    for container in ax.containers:
        for bar in container:
            height = bar.get_height()
            if height > 0:
                x = bar.get_x() + bar.get_width() / 2
                y = bar.get_y() + height / 2
                ax.text(x, y, f"{int(height)}", ha='center', va='center', fontsize=8, color="black")
    plt.savefig(STATUS_CHART_PATH)
    plt.close()
    st.session_state.chart_generated = True

def generate_violations(report_df, batch_size):
    violations_df = pd.DataFrame(columns=["State", "LGA", "Ward", "School ID", "School Name", "Dataset UID", "datasets_name", "Period", "Validation Rule Name", "Left Side Value", "Right Side Value", "Importance"])
    no_upload_rows = report_df[report_df["Status"].str.startswith("⚠️ Logged in, no data upload")]
    
    if not no_upload_rows.empty:
        st.write(f"Processing {len(no_upload_rows)} rows with ⚠️ status")
        
        # Deduplicate by School ID and collect all dataset_uids
        unique_ous = no_upload_rows.groupby("School ID")["datasetuid"].apply(list).reset_index()
        batches = [unique_ous.iloc[i:i + batch_size][["School ID", "datasetuid"]].values.tolist() for i in range(0, len(unique_ous), batch_size)]
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        with ThreadPoolExecutor(max_workers=4) as executor:
            future_to_batch = {executor.submit(get_validation_violations_batch, batch, report_df): batch for batch in batches}
            for i, future in enumerate(as_completed(future_to_batch)):
                batch = future_to_batch[future]
                progress = (i + 1) / len(batches)
                progress_bar.progress(progress)
                status_text.text(f"Processing batch {i+1}/{len(batches)}")
                
                try:
                    results = future.result()
                    for (orgunit_id, dataset_uids), result in results.items():
                        violations = result["violations"]
                        state = result["state"]
                        lga = result["lga"]
                        ward = result["ward"]
                        dataset_name = result["dataset_name"]
                        orgunit_name = get_orgunit_name(orgunit_id)
                        
                        if violations:
                            for v in violations:
                                rule = v.get("validationRule", {})
                                rule_name = rule.get("name", "Unnamed Rule")
                                violations_df = pd.concat([violations_df, pd.DataFrame([{
                                    "State": state,
                                    "LGA": lga,
                                    "Ward": ward,
                                    "School ID": orgunit_id,
                                    "School Name": orgunit_name,
                                    "Dataset UID": v.get("dataSet", {}).get("id", dataset_uids[0]),
                                    "datasets_name": dataset_name,
                                    "Period": v.get("period", {}).get("name", "2024"),
                                    "Validation Rule Name": rule_name,
                                    "Left Side Value": v.get("leftsideValue"),
                                    "Right Side Value": v.get("rightsideValue"),
                                    "Importance": rule.get("importance", "N/A")
                                }])], ignore_index=True)
                except Exception as e:
                    st.error(f"Batch processing failed: {e}")
        
        progress_bar.empty()
        status_text.empty()
    
    # Update Excel file with violations
    if not violations_df.empty:
        temp_excel = os.path.join(OUTPUT_FOLDER, "temp_violations.xlsx")
        try:
            if os.path.exists(REPORT_EXCEL):
                shutil.copy(REPORT_EXCEL, temp_excel)
                wb = load_workbook(temp_excel)
            else:
                report_df.to_excel(temp_excel, index=False, sheet_name="Report")
                wb = load_workbook(temp_excel)
                ws_report = wb["Report"]
                for cell in ws_report[1]:
                    cell.font = Font(bold=True)
                for column_cells in ws_report.columns:
                    max_length = max(len(str(cell.value or "")) for cell in column_cells)
                    ws_report.column_dimensions[column_cells[0].column_letter].width = max_length + 2
                status_colors = {
                    "✅ Logged in and uploaded data": "C6EFCE",
                    "⚠️ Logged in, no data upload": "FFF2CC",
                    "❌ User exists, yet to login": "F8CBAD",
                    "🚫 No user account": "D9D9D9"
                }
                status_col = [cell.value for cell in ws_report[1]].index("Status") + 1
                for row in ws_report.iter_rows(min_row=2, max_row=ws_report.max_row):
                    status = row[status_col - 1].value
                    if isinstance(status, str):
                        for key, color in status_colors.items():
                            if status.startswith(key):
                                fill = PatternFill(start_color=color, fill_type="solid")
                                for cell in row:
                                    cell.fill = fill
                                break
            
            if "Violations" in wb.sheetnames:
                ws_violations = wb["Violations"]
                ws_violations.delete_rows(2, ws_violations.max_row)  # Clear existing data
            else:
                ws_violations = wb.create_sheet("Violations")
                headers = ["State", "LGA", "Ward", "School ID", "School Name", "Dataset UID", "datasets_name", "Period", "Validation Rule Name", "Left Side Value", "Right Side Value", "Importance"]
                ws_violations.append(headers)
                for cell in ws_violations[1]:
                    cell.font = Font(bold=True)
            
            for _, row in violations_df.iterrows():
                ws_violations.append([row[col] for col in headers])
            
            for column_cells in ws_violations.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                ws_violations.column_dimensions[column_cells[0].column_letter].width = max_length + 2
            
            wb.save(temp_excel)
            if os.path.exists(REPORT_EXCEL):
                os.remove(REPORT_EXCEL)
            shutil.move(temp_excel, REPORT_EXCEL)
        except Exception as e:
            st.error(f"Failed to update Excel file with violations: {e}")
            if os.path.exists(temp_excel):
                os.remove(temp_excel)
    
    return violations_df

def get_quick_stats(report_df, violations_df):
    if report_df.empty and violations_df.empty:
        return {
            "total_schools": 0,
            "completion_rate": 0,
            "active_users": 0,
            "pending_uploads": 0,
            "fully_completed_datasets": 0,
            "users_no_login": 0,
            "users_no_access": 0,
            "all_users_with_login": 0,
            "users_with_uploads": 0,
            "total_violations": 0
        }

    current_date = datetime.now()
    reporting_period_start = (current_date - timedelta(days=30)).strftime('%Y-%m-%d')

    unique_schools = len(report_df["School ID"].unique())
    completed_schools = len(report_df[report_df["Status"].str.startswith("✅")])
    completion_rate = round((completed_schools / unique_schools) * 100, 1) if unique_schools > 0 else 0
    
    active_users = report_df[
        (report_df["Last Login"].notna()) & 
        (report_df["Last Login"] >= reporting_period_start)
    ]["Usernames"].str.split(',').explode().str.strip().nunique()
    
    pending_uploads = len(report_df[report_df["Status"].str.startswith("⚠️")])
    users_no_login = report_df[
        report_df["Status"].str.startswith("❌")
    ]["User Count"].sum()
    
    users_no_access = len(report_df[
        report_df["Status"].str.startswith("🚫")
    ])
    
    all_users_with_login = report_df[
        report_df["Status"].str.startswith("✅") | 
        report_df["Status"].str.startswith("⚠️")
    ]["User Count"].sum()
    
    users_with_uploads = report_df[
        report_df["Status"].str.startswith("✅")
    ]["User Count"].sum()
    
    dataset_completion = report_df.groupby("datasets_name")["Status"].apply(
        lambda x: (x.str.startswith("✅").sum() / len(x) * 100)
    ).to_dict()
    
    fully_completed_datasets = sum(1 for v in dataset_completion.values() if v == 100)
    total_violations = len(violations_df)

    return {
        "total_schools": unique_schools,
        "completion_rate": completion_rate,
        "active_users": active_users,
        "pending_uploads": pending_uploads,
        "fully_completed_datasets": fully_completed_datasets,
        "users_no_login": users_no_login,
        "users_no_access": users_no_access,
        "all_users_with_login": all_users_with_login,
        "users_with_uploads": users_with_uploads,
        "total_violations": total_violations
    }

def get_violation_stats(violations_df):
    if violations_df.empty:
        return []
    
    violation_counts = violations_df["School ID"].value_counts().head(5).to_dict()
    return [
        {
            "school_id": school_id,
            "count": count,
            "school_name": violations_df[violations_df["School ID"] == school_id]["School Name"].iloc[0]
        }
        for school_id, count in violation_counts.items()
    ]

def get_download_link(file_path, file_label):
    if os.path.exists(file_path):
        with open(file_path, "rb") as f:
            data = f.read()
        b64 = base64.b64encode(data).decode()
        href = f'<a href="data:application/octet-stream;base64,{b64}" download="{os.path.basename(file_path)}">{file_label}</a>'
        return href
    return None

def apply_advanced_filters(data, filters):
    if not isinstance(data, pd.DataFrame) or data.empty:
        return data
    
    filtered = data.copy()
    
    # Status filter
    if filters['status']:
        status_conditions = []
        if "✅" in filters['status']:
            status_conditions.append(filtered["Status"].str.startswith("✅"))
        if "⚠️" in filters['status']:
            status_conditions.append(filtered["Status"].str.startswith("⚠️"))
        if "❌" in filters['status']:
            status_conditions.append(filtered["Status"].str.startswith("❌"))
        if "🚫" in filters['status']:
            status_conditions.append(filtered["Status"].str.startswith("🚫"))
        
        if status_conditions:
            filtered = filtered[pd.concat(status_conditions, axis=1).any(axis=1)]
    
    # Dataset filter
    if filters['dataset']:
        filtered = filtered[filtered["datasets_name"].isin(filters['dataset'])]
    
    # User count range
    if filters['user_count_min']:
        filtered = filtered[filtered["User Count"] >= int(filters['user_count_min'])]
    if filters['user_count_max']:
        filtered = filtered[filtered["User Count"] <= int(filters['user_count_max'])]
    
    # Last login date range
    if filters['last_login_start']:
        filtered = filtered[
            (filtered["Last Login"] >= filters['last_login_start']) | 
            (filtered["Last Login"].isna())
        ]
    if filters['last_login_end']:
        filtered = filtered[
            (filtered["Last Login"] <= filters['last_login_end']) | 
            (filtered["Last Login"].isna())
        ]
    
    return filtered

def prepare_chart_data(report_df):
    if report_df.empty:
        return []
    
    status_count = report_df["Status"].apply(
        lambda s: "Completed" if s.startswith("✅") else
                  "Logged In, No Upload" if s.startswith("⚠️") else
                  "No Login" if s.startswith("❌") else
                  "No User"
    ).value_counts().to_dict()
    
    return [
        {
            "name": "Completed",
            "value": status_count.get("Completed", 0),
            "color": "#4CAF50"
        },
        {
            "name": "Logged In, No Upload",
            "value": status_count.get("Logged In, No Upload", 0),
            "color": "#FFC107"
        },
        {
            "name": "No Login",
            "value": status_count.get("No Login", 0),
            "color": "#F44336"
        },
        {
            "name": "No User",
            "value": status_count.get("No User", 0),
            "color": "#9E9E9E"
        }
    ]

# Streamlit UI
st.set_page_config(layout="wide", page_title="DHIS2 EMIS Dashboard", page_icon="📊")

st.title("DHIS2 EMIS Data Upload Dashboard")

# Initialize session state
if 'selected_state' not in st.session_state:
    st.session_state.selected_state = STATE_OPTIONS[0]['uid']
if 'show_tables' not in st.session_state:
    st.session_state.show_tables = True
if 'chart_type' not in st.session_state:
    st.session_state.chart_type = 'bar'

# State selection and data fetching
col1, col2, col3, col4 = st.columns(4)
with col1:
    selected_state = st.selectbox(
        "Select State",
        options=STATE_OPTIONS,
        format_func=lambda x: x['name'],
        index=0,
        key='state_select'
    )
    st.session_state.selected_state = selected_state['uid']

with col2:
    batch_size = st.selectbox(
        "Batch Size",
        options=[5, 10, 15],
        index=1,
        key='batch_size_select'
    )

with col3:
    if st.button("Fetch Datasets", key='fetch_datasets'):
        with st.spinner("Fetching datasets..."):
            st.session_state.raw_report_data = generate_report(selected_state['uid'])
            if st.session_state.raw_report_data is not None:
                st.success("Datasets fetched successfully!")

with col4:
    if st.button("Generate Full Report", key='generate_report'):
        with st.spinner("Generating full report..."):
            st.session_state.report_df = generate_full_report()
            if st.session_state.report_df is not None:
                generate_chart(st.session_state.report_df)
                st.success("Full report generated successfully!")

# Check violations button
if st.session_state.report_df is not None:
    if st.button("Check Violations", key='fetch_violations'):
        with st.spinner("Checking for violations..."):
            st.session_state.violations_df = generate_violations(st.session_state.report_df, batch_size)
            if st.session_state.violations_df is not None:
                st.success(f"Found {len(st.session_state.violations_df)} violations")

# Auto-refresh settings
with st.expander("Auto Refresh Settings"):
    auto_refresh = st.checkbox("Enable Auto Refresh", value=False, key='auto_refresh')
    refresh_interval = st.number_input(
        "Refresh Interval (seconds)", 
        min_value=60, 
        max_value=3600, 
        value=300, 
        step=60,
        key='refresh_interval',
        disabled=not auto_refresh
    )

# Filter controls
st.subheader("Filters")
filter_col1, filter_col2 = st.columns(2)
with filter_col1:
    filter_lga = st.text_input("Filter by LGA", key='filter_lga')
with filter_col2:
    filter_org_unit = st.text_input("Filter by School Name", key='filter_org_unit')

# Advanced filters
with st.expander("Advanced Filters"):
    status_options = [
        {"value": "✅", "label": "Logged in and uploaded data"},
        {"value": "⚠️", "label": "Logged in, no data upload"},
        {"value": "❌", "label": "User exists, yet to login"},
        {"value": "🚫", "label": "No user account"}
    ]
    
    selected_status = st.multiselect(
        "Status",
        options=status_options,
        format_func=lambda x: x['label'],
        default=[],
        key='status_filter'
    )
    st.session_state.advanced_filters['status'] = [s['value'] for s in selected_status]
    
    if st.session_state.report_df is not None:
        dataset_options = st.session_state.report_df['datasets_name'].unique()
        selected_datasets = st.multiselect(
            "Datasets",
            options=dataset_options,
            default=[],
            key='dataset_filter'
        )
        st.session_state.advanced_filters['dataset'] = selected_datasets
    
    user_count_col1, user_count_col2 = st.columns(2)
    with user_count_col1:
        user_count_min = st.number_input(
            "Min User Count",
            min_value=0,
            value=0,
            key='user_count_min'
        )
        st.session_state.advanced_filters['user_count_min'] = user_count_min if user_count_min > 0 else ''
    
    with user_count_col2:
        user_count_max = st.number_input(
            "Max User Count",
            min_value=0,
            value=0,
            key='user_count_max'
        )
        st.session_state.advanced_filters['user_count_max'] = user_count_max if user_count_max > 0 else ''
    
    login_date_col1, login_date_col2 = st.columns(2)
    with login_date_col1:
        last_login_start = st.date_input(
            "Last Login After",
            value=None,
            key='last_login_start'
        )
        st.session_state.advanced_filters['last_login_start'] = last_login_start.strftime('%Y-%m-%d') if last_login_start else ''
    
    with login_date_col2:
        last_login_end = st.date_input(
            "Last Login Before",
            value=None,
            key='last_login_end'
        )
        st.session_state.advanced_filters['last_login_end'] = last_login_end.strftime('%Y-%m-%d') if last_login_end else ''

# Apply filters to data
filtered_report = pd.DataFrame()
filtered_violations = pd.DataFrame()

if st.session_state.report_df is not None:
    filtered_report = st.session_state.report_df.copy()
    
    # Apply basic filters
    if filter_lga:
        filtered_report = filtered_report[filtered_report['LGA'].str.contains(filter_lga, case=False, na=False)]
    if filter_org_unit:
        filtered_report = filtered_report[filtered_report['School Name'].str.contains(filter_org_unit, case=False, na=False)]
    
    # Apply advanced filters
    filtered_report = apply_advanced_filters(filtered_report, st.session_state.advanced_filters)

if st.session_state.violations_df is not None:
    filtered_violations = st.session_state.violations_df.copy()
    if filter_lga:
        filtered_violations = filtered_violations[filtered_violations['LGA'].str.contains(filter_lga, case=False, na=False)]
    if filter_org_unit:
        filtered_violations = filtered_violations[filtered_violations['School Name'].str.contains(filter_org_unit, case=False, na=False)]

# Quick statistics
quick_stats = get_quick_stats(filtered_report, filtered_violations)
violation_stats = get_violation_stats(filtered_violations)

# Display statistics
st.subheader("Key Metrics")
metric_cols = st.columns(5)
with metric_cols[0]:
    st.metric("Total Schools", quick_stats['total_schools'])
with metric_cols[1]:
    st.metric("Completion Rate", f"{quick_stats['completion_rate']}%")
with metric_cols[2]:
    st.metric("Active Users", quick_stats['active_users'])
with metric_cols[3]:
    st.metric("Pending Uploads", quick_stats['pending_uploads'])
with metric_cols[4]:
    st.metric("Validation Violations", quick_stats['total_violations'])

# More detailed statistics
st.subheader("Detailed Statistics")
stats_col1, stats_col2, stats_col3, stats_col4 = st.columns(4)
with stats_col1:
    st.metric("Fully Completed Datasets", quick_stats['fully_completed_datasets'])
with stats_col2:
    st.metric("Users Without Login", quick_stats['users_no_login'])
with stats_col3:
    st.metric("Schools Without Users", quick_stats['users_no_access'])
with stats_col4:
    st.metric("Users With Uploads", quick_stats['users_with_uploads'])

# Charts
if st.session_state.report_df is not None:
    st.subheader("Data Visualization")
    
    chart_tab1, chart_tab2 = st.tabs(["Status Distribution", "LGA Performance"])
    
    with chart_tab1:
        chart_type = st.selectbox(
            "Chart Type",
            options=['bar', 'pie', 'line'],
            index=0,
            key='chart_type_select'
        )
        
        chart_data = prepare_chart_data(filtered_report)
        
        if chart_data:
            fig, ax = plt.subplots(figsize=(10, 6))
            
            if chart_type == 'bar':
                df = pd.DataFrame(chart_data)
                sns.barplot(data=df, x='name', y='value', palette=[d['color'] for d in chart_data], ax=ax)
                ax.set_title("Upload Status Distribution")
                ax.set_xlabel("Status")
                ax.set_ylabel("Count")
                plt.xticks(rotation=45)
                
            elif chart_type == 'pie':
                df = pd.DataFrame(chart_data)
                ax.pie(
                    df['value'],
                    labels=df['name'],
                    colors=[d['color'] for d in chart_data],
                    autopct='%1.1f%%',
                    startangle=90
                )
                ax.set_title("Upload Status Distribution")
                ax.axis('equal')
                
            elif chart_type == 'line':
                df = pd.DataFrame(chart_data)
                sns.lineplot(data=df, x='name', y='value', marker='o', color='#8884d8', ax=ax)
                ax.set_title("Upload Status Distribution")
                ax.set_xlabel("Status")
                ax.set_ylabel("Count")
                plt.xticks(rotation=45)
            
            st.pyplot(fig)
    
    with chart_tab2:
        if not filtered_report.empty:
            lga_stats = filtered_report.groupby('LGA').agg(
                Total=('Status', 'count'),
                Completed=('Status', lambda x: sum(x.str.startswith("✅"))),
            ).reset_index()
            lga_stats['Completion Rate'] = (lga_stats['Completed'] / lga_stats['Total'] * 100).round(1)
            
            top_lgas = lga_stats.nlargest(10, 'Completion Rate')
            
            fig, ax = plt.subplots(figsize=(12, 6))
            sns.barplot(data=top_lgas, x='LGA', y='Completion Rate', palette='viridis', ax=ax)
            ax.set_title("Top 10 LGAs by Completion Rate")
            ax.set_xlabel("LGA")
            ax.set_ylabel("Completion Rate (%)")
            plt.xticks(rotation=45)
            st.pyplot(fig)

# Top violations
if violation_stats:
    st.subheader("Top Schools with Validation Violations")
    cols = st.columns(len(violation_stats))
    for i, stat in enumerate(violation_stats):
        with cols[i]:
            st.metric(
                label=f"{stat['school_name']}",
                value=stat['count'],
                help=f"School ID: {stat['school_id']}"
            )

# Data tables
show_tables = st.checkbox("Show Data Tables", value=True, key='show_tables_checkbox')

if show_tables and st.session_state.report_df is not None:
    st.subheader("School Report Data")
    st.dataframe(
        filtered_report,
        use_container_width=True,
        height=400,
        column_config={
            "Status": st.column_config.TextColumn(
                "Status",
                help="Status of data upload",
                width="medium"
            ),
            "User Count": st.column_config.NumberColumn(
                "User Count",
                help="Number of users associated",
                width="small"
            )
        }
    )

if show_tables and st.session_state.violations_df is not None and not filtered_violations.empty:
    st.subheader("Validation Violations")
    st.dataframe(
        filtered_violations,
        use_container_width=True,
        height=400
    )

# Export options
with st.expander("Export Data"):
    export_col1, export_col2, export_col3 = st.columns(3)
    
    with export_col1:
        export_format = st.selectbox(
            "Format",
            options=["excel", "csv", "json"],
            key='export_format'
        )
        st.session_state.export_options['format'] = export_format
    
    with export_col2:
        include_violations = st.checkbox(
            "Include Violations",
            value=True,
            key='include_violations'
        )
        st.session_state.export_options['include_violations'] = include_violations
    
    with export_col3:
        st.write("")  # Spacer
        if st.button("Generate Export", key='generate_export'):
            with st.spinner("Preparing export..."):
                # Prepare data for export
                export_data = filtered_report.copy()
                
                if include_violations and st.session_state.violations_df is not None:
                    export_data = pd.concat([export_data, filtered_violations], axis=0)
                
                # Select only the chosen columns
                if st.session_state.export_options['selected_columns']:
                    export_data = export_data[st.session_state.export_options['selected_columns']]
                
                # Create download link based on format
                if export_format == 'excel':
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        export_data.to_excel(writer, index=False, sheet_name='Report')
                        if include_violations and not filtered_violations.empty:
                            filtered_violations.to_excel(writer, index=False, sheet_name='Violations')
                    output.seek(0)
                    st.download_button(
                        label="Download Excel",
                        data=output,
                        file_name=f"emis_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                elif export_format == 'csv':
                    csv = export_data.to_csv(index=False)
                    st.download_button(
                        label="Download CSV",
                        data=csv,
                        file_name=f"emis_export_{datetime.now().strftime('%Y%m%d')}.csv",
                        mime="text/csv"
                    )
                
                elif export_format == 'json':
                    json_data = export_data.to_json(orient='records')
                    st.download_button(
                        label="Download JSON",
                        data=json_data,
                        file_name=f"emis_export_{datetime.now().strftime('%Y%m%d')}.json",
                        mime="application/json"
                    )

# Column selection for export
if st.session_state.report_df is not None:
    st.write("Select columns to include in export:")
    col_options = st.session_state.report_df.columns.tolist()
    selected_cols = st.multiselect(
        "Columns",
        options=col_options,
        default=["State", "LGA", "Ward", "School ID", "School Name", "datasets_name", "Status", "User Count", "Usernames", "Last Login"],
        key='export_columns'
    )
    st.session_state.export_options['selected_columns'] = selected_cols

# Download links
st.subheader("Download Reports")
if os.path.exists(REPORT_EXCEL):
    with open(REPORT_EXCEL, "rb") as f:
        st.download_button(
            label="Download Full Report (Excel)",
            data=f,
            file_name="State_school_user_upload_status_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

if st.session_state.chart_generated and os.path.exists(STATUS_CHART_PATH):
    with open(STATUS_CHART_PATH, "rb") as f:
        st.download_button(
            label="Download Status Chart (PNG)",
            data=f,
            file_name="LGA_Status_Chart.png",
            mime="image/png"
        )

# Notifications
if st.session_state.notifications:
    with st.expander("Notifications", expanded=True):
        for notification in st.session_state.notifications[-3:]:  # Show last 3 notifications
            if notification['type'] == 'success':
                st.success(notification['message'])
            elif notification['type'] == 'error':
                st.error(notification['message'])
            else:
                st.info(notification['message'])

# Auto-refresh logic
if auto_refresh:
    time.sleep(refresh_interval)
    st.rerun()