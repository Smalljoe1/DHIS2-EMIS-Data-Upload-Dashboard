from flask import Flask, jsonify, send_file, send_from_directory, request
from flask_cors import CORS
import pandas as pd
import requests
from io import StringIO
import os
from dotenv import load_dotenv
import csv
import matplotlib
matplotlib.use('Agg')  # Use Agg backend for non-GUI rendering
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
import shutil

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})

# Configuration
BASE_URL = "https://emis.dhis2nigeria.org.ng/dhis/api"
HEADERS = {
    "Content-Type": "application/json",
    "Authorization": f"ApiToken {os.getenv('DHIS2_API_TOKEN', 'd2pat_DAULytLK7GUBSlfknZ6ffVxIfifC8T9c3885349470')}"
}
DATASET_UIDS = [
    "MLTLNUmvS8r", "uSw8GwPO417", "W36yBpVEUkH",
    "pJydop5Fpsz", "XERITHzkeSI", "RlfDdEEZ317"
]
PERIOD = "LAST_YEAR"
OUTPUT_FOLDER = "./data"
MAPPING_FILE = os.path.join(OUTPUT_FOLDER, "dataset_uid_mapping.csv")
COMBINED_FILE = os.path.join(OUTPUT_FOLDER, "dataset_completion_report_EMIS.csv")
USER_FILE = os.path.join(OUTPUT_FOLDER, "dhis2_users.csv")
REPORT_EXCEL = os.path.join(OUTPUT_FOLDER, "State_school_user_upload_status_report.xlsx")
NO_UPLOAD_FILE = os.path.join(OUTPUT_FOLDER, "logged_in_no_upload.xlsx")
STATUS_CHART_PATH = os.path.join(OUTPUT_FOLDER, "LGA_Status_Chart.png")

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Load UID-to-Name Mapping
try:
    uid_map_df = pd.read_csv(MAPPING_FILE, dtype=str)
    if not {'dataset_uids', 'dataset_names'}.issubset(uid_map_df.columns):
        raise ValueError("Mapping file must contain 'dataset_uids' and 'dataset_names' columns.")
except FileNotFoundError:
    uid_map_df = pd.DataFrame({'dataset_uids': DATASET_UIDS, 'dataset_names': ['Dataset ' + uid for uid in DATASET_UIDS]})
    uid_map_df.to_csv(MAPPING_FILE, index=False)

# Fetch dataset report
def fetch_dataset_report(uid, org_units):
    params = {
        "dimension": [f"dx:{uid}.ACTUAL_REPORTS;{uid}.EXPECTED_REPORTS", f"ou:{org_units}"],
        "filter": f"pe:{PERIOD}",
        "tableLayout": "true",
        "columns": "dx",
        "rows": "ou",
        "skipRounding": "false",
        "completedOnly": "false",
        "hideEmptyRows": "true",
        "showHierarchy": "true"
    }
    try:
        response = requests.get(f"{BASE_URL}/analytics.csv", headers=HEADERS, params=params, timeout=120)
        response.raise_for_status()
        df = pd.read_csv(StringIO(response.text))
        df["datasetuid"] = uid
        dataset_name = uid_map_df.loc[uid_map_df['dataset_uids'] == uid, 'dataset_names'].values[0]
        df["datasets_name"] = dataset_name
        return df
    except requests.RequestException:
        return None

# Fetch all datasets
@app.route('/api/datasets', methods=['GET'])
def get_datasets():
    org_units = request.args.get('orgUnits', 'LEVEL-st3hrLkzuMb;Ym1fEhWFWYI')  # Default to Kano state
    all_dfs = []
    for uid in DATASET_UIDS:
        df = fetch_dataset_report(uid, org_units)
        if df is not None and not df.empty:
            all_dfs.append(df)
    if all_dfs:
        final_df = pd.concat(all_dfs, ignore_index=True)
        final_df.to_csv(COMBINED_FILE, index=False)
        return jsonify(final_df.to_dict(orient='records'))
    return jsonify({"error": "No datasets fetched"}), 500

# Fetch users
@app.route('/api/users', methods=['GET'])
def get_users():
    # state_uid = request.args.get('orgUnits', 'LEVEL-st3hrLkzuMb;Ym1fEhWFWYI')  # Default to Yobe State
    url = f"{BASE_URL}/users.json?fields=id,name,username,userGroups[name],userRoles[name],lastLogin,organisationUnits[ancestors[name],name,id,level]&paging=true&pageSize=10000"
    seen_ids = set()
    users = []
    while url:
        response = requests.get(url, headers=HEADERS, timeout=60)
        if response.status_code != 200:
            return jsonify({"error": "Failed to fetch users"}), 500
        data = response.json()
        for user in data.get("users", []):
            uid = user.get("id")
            if not uid or uid in seen_ids:
                continue
            seen_ids.add(uid)
            user_groups = ", ".join([g["name"] for g in user.get("userGroups", [])])
            user_roles = ", ".join([r["name"] for r in user.get("userRoles", [])])
            school_paths = []
            school_uids = []
            for ou in user.get("organisationUnits", []):
                if ou.get("level") == 5:  # Filter for level 5 (schools)
                    path = " > ".join([a["name"] for a in ou.get("ancestors", [])] + [ou["name"]])
                    school_paths.append(path)
                    school_uids.append(ou.get("id", ""))
            if school_uids:  # Only include users with level 5 OUs
                users.append({
                    "id": uid,
                    "name": user.get("name", ""),
                    "username": user.get("username", ""),
                    "userGroups": user_groups,
                    "userRoles": user_roles,
                    "lastLogin": user.get("lastLogin", ""),
                    "schoolPath": ", ".join(school_paths),
                    "schoolUID": ", ".join(school_uids)
                })
        url = data.get("pager", {}).get("nextPage")
    pd.DataFrame(users).to_csv(USER_FILE, index=False)
    return jsonify(users)

# Validation Functions
def get_orgunit_name(orgunit_id):
    url = f"{BASE_URL}/33/organisationUnits/{orgunit_id}.json?fields=name"
    try:
        res = requests.get(url, headers=HEADERS, timeout=10)
        res.raise_for_status()
        return res.json().get("name", "Unknown")
    except:
        return "Unknown"

def get_validation_violations_batch(batch, report_df):
    results = {}
    for orgunit_id, dataset_uids in batch:
        orgunit_name = get_orgunit_name(orgunit_id)
        print(f"\n🔍 Checking: {orgunit_name} ({orgunit_id}) - Datasets: {dataset_uids}")
        violations = []
        # Find matching row in report_df to get State, LGA, Ward, and datasets_name
        matching_row = report_df[report_df["School ID"] == orgunit_id].iloc[0] if not report_df.empty and orgunit_id in report_df["School ID"].values else None
        state = matching_row["State"] if matching_row is not None else "Unknown"
        lga = matching_row["LGA"] if matching_row is not None else "Unknown"
        ward = matching_row["Ward"] if matching_row is not None else "Unknown"
        dataset_name = matching_row["datasets_name"] if matching_row is not None else "Unknown"
        
        for dataset_uid in dataset_uids:
            url = f"{BASE_URL}/33/validation/dataSet/{dataset_uid}.json?pe=2024&ou={orgunit_id}"
            try:
                res = requests.get(url, headers=HEADERS, timeout=15)
                res.raise_for_status()
                violations.extend(res.json().get("validationRuleViolations", []))
            except requests.exceptions.RequestException as e:
                print(f"   ⚠️ Failed for {orgunit_id} with {dataset_uid}: {e}")
        if not violations:
            print("   ℹ️ No violations found.")
        else:
            print(f"   ⚠️ {len(violations)} violation(s) found:")
            for v in violations:
                rule = v.get("validationRule", {})
                rule_name = rule.get("name", "Unnamed Rule")
                print(f"     - {rule_name}")
        results[(orgunit_id, tuple(dataset_uids))] = {
            "violations": violations,
            "state": state,
            "lga": lga,
            "ward": ward,
            "dataset_name": dataset_name
        }
    return results

# Generate report (without violations)
@app.route('/api/report', methods=['GET'])
def get_report():
    users_df = pd.read_csv(USER_FILE, dtype=str).fillna("")
    dataset_df = pd.read_csv(COMBINED_FILE, dtype=str).fillna("")
    
    users_df["schoolUID_original"] = users_df["schoolUID"]
    dataset_df["organisationunitid_original"] = dataset_df["organisationunitid"]
    users_df["schoolUID"] = users_df["schoolUID"].str.strip().str.lower().str.replace(r'[^a-z0-9]', '', regex=True)
    dataset_df["organisationunitid"] = dataset_df["organisationunitid"].str.strip().str.lower().str.replace(r'[^a-z0-9]', '', regex=True)
    
    actual_cols = [col for col in dataset_df.columns if "Actual reports" in col]
    dataset_df["Total_Actual"] = dataset_df[actual_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1)
    
    uploaded_schools = set(dataset_df[dataset_df["Total_Actual"] > 0]["organisationunitid"])
    all_schools = set(dataset_df["organisationunitid"])
    
    report_rows = []
    for school_id in sorted(all_schools):
        school_data = dataset_df[dataset_df["organisationunitid"] == school_id].iloc[0]
        school_name = school_data["organisationunitname"]
        original_school_id = school_data["organisationunitid_original"]
        dataset_uid = school_data.get("datasetuid", "")
        dataset_name = school_data.get("datasets_name", "")
        
        school_users = users_df[users_df["schoolUID"] == school_id].copy()
        school_users["schoolUID"] = school_users["schoolUID_original"]
        
        if school_users.empty:
            state = school_data.get("orgunitlevel2", "")
            lga = school_data.get("orgunitlevel3", "")
            ward = school_data.get("orgunitlevel4", "")
            status = "🚫 No user account"
            last_login = ""
        else:
            first_user = school_users.iloc[0]
            state = school_data.get("orgunitlevel2", "")
            lga = school_data.get("orgunitlevel3", "")
            ward = school_data.get("orgunitlevel4", "")
            has_login = school_users["lastLogin"].str.strip().ne("").any()
            has_upload = school_id in uploaded_schools
            if has_upload:
                status = "✅ Logged in and uploaded data"
            elif has_login:
                status = "⚠️ Logged in, no data upload"
            else:
                status = "❌ User exists, yet to login"
            if len(school_users) > 1:
                status += f" (👥 {len(school_users)} users)"
            last_logins = pd.to_datetime(school_users["lastLogin"], errors='coerce', utc=True).dropna()
            last_login = last_logins.max().strftime('%Y-%m-%d %H:%M:%S UTC') if not last_logins.empty else ""
        
        report_rows.append({
            "State": state,
            "LGA": lga,
            "Ward": ward,
            "School ID": original_school_id,
            "School Name": school_name,
            "datasetuid": dataset_uid,
            "datasets_name": dataset_name,
            "Status": status,
            "User Count": len(school_users),
            "Usernames": ", ".join(school_users["username"]),
            "Last Login": last_login
        })
    
    report_df = pd.DataFrame(report_rows)
    ordered_cols = ["State", "LGA", "Ward", "School ID", "School Name", "datasetuid", "datasets_name", "Status", "User Count", "Usernames", "Last Login"]
    report_df = report_df[ordered_cols]
    
    # Handle Excel file creation with corruption check
    temp_excel = os.path.join(OUTPUT_FOLDER, "temp_report.xlsx")
    try:
        # Initial write to a temporary file
        report_df.to_excel(temp_excel, index=False, sheet_name="Report")
        
        # Format Excel
        wb = load_workbook(temp_excel)
        ws_report = wb["Report"]
        for cell in ws_report[1]:
            cell.font = Font(bold=True)
        for column_cells in ws_report.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws_report.column_dimensions[column_cells[0].column_letter].width = max_length + 2
        status_colors = {
            "✅ Logged in and uploaded data": "C6EFCE",
            "⚠️ Logged in, no data upload": "FFF2CC",
            "❌ User exists, yet to login": "F8CBAD",
            "🚫 No user account": "D9D9D9"
        }
        status_col = [cell.value for cell in ws_report[1]].index("Status") + 1
        for row in ws_report.iter_rows(min_row=2, max_row=ws_report.max_row):
            status = row[status_col - 1].value
            if isinstance(status, str):
                for key, color in status_colors.items():
                    if status.startswith(key):
                        fill = PatternFill(start_color=color, fill_type="solid")
                        for cell in row:
                            cell.fill = fill
                        break
        
        wb.save(temp_excel)
        
        # Replace the original file only if the temp file is valid
        if os.path.exists(REPORT_EXCEL):
            os.remove(REPORT_EXCEL)
        shutil.move(temp_excel, REPORT_EXCEL)
    except Exception as e:
        print(f"⚠️ Failed to save Excel file: {e}")
        if os.path.exists(temp_excel):
            os.remove(temp_excel)
        raise

    # Generate chart
    report_df["Clean Status"] = report_df["Status"].apply(
        lambda s: "Logged in & uploaded" if s.startswith("✅") else
                  "Logged in, no upload" if s.startswith("⚠️") else
                  "User exists, no login" if s.startswith("❌") else
                  "No user account"
    )
    status_counts = report_df.groupby(["LGA", "Clean Status"]).size().unstack(fill_value=0)
    sorted_lgas = status_counts.sum(axis=1).sort_values(ascending=False)
    top_lgas = sorted_lgas.head(20).index if len(sorted_lgas) > 20 else sorted_lgas.index
    status_counts = status_counts.loc[top_lgas]
    
    fig, ax = plt.subplots(figsize=(16, 9))
    status_counts.plot(kind='bar', stacked=True, colormap='Set2', ax=ax)
    num_lgas = len(status_counts)
    title = f"All LGAs by School Upload & Login Status ({num_lgas} total)" if num_lgas < 20 else "Top 20 LGAs by School Upload & Login Status"
    plt.title(title, fontsize=16)
    plt.xlabel("LGA", fontsize=12)
    plt.ylabel("Number of Schools", fontsize=12)
    plt.xticks(rotation=30, ha='right')
    plt.legend(title="Status", bbox_to_anchor=(1.02, 1), loc='upper left')
    plt.tight_layout()
    for container in ax.containers:
        for bar in container:
            height = bar.get_height()
            if height > 0:
                x = bar.get_x() + bar.get_width() / 2
                y = bar.get_y() + height / 2
                ax.text(x, y, f"{int(height)}", ha='center', va='center', fontsize=8, color="black")
    plt.savefig(STATUS_CHART_PATH)
    plt.close()
    
    # Return report without violations
    return jsonify({"report": report_df.to_dict(orient='records'), "violations": []})

# Generate violations report and update Excel
@app.route('/api/violations', methods=['GET'])
def get_violations():
    users_df = pd.read_csv(USER_FILE, dtype=str).fillna("")
    dataset_df = pd.read_csv(COMBINED_FILE, dtype=str).fillna("")
    
    users_df["schoolUID_original"] = users_df["schoolUID"]
    dataset_df["organisationunitid_original"] = dataset_df["organisationunitid"]
    users_df["schoolUID"] = users_df["schoolUID"].str.strip().str.lower().str.replace(r'[^a-z0-9]', '', regex=True)
    dataset_df["organisationunitid"] = dataset_df["organisationunitid"].str.strip().str.lower().str.replace(r'[^a-z0-9]', '', regex=True)
    
    actual_cols = [col for col in dataset_df.columns if "Actual reports" in col]
    dataset_df["Total_Actual"] = dataset_df[actual_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1)
    
    uploaded_schools = set(dataset_df[dataset_df["Total_Actual"] > 0]["organisationunitid"])
    all_schools = set(dataset_df["organisationunitid"])
    
    report_rows = []
    for school_id in sorted(all_schools):
        school_data = dataset_df[dataset_df["organisationunitid"] == school_id].iloc[0]
        school_name = school_data["organisationunitname"]
        original_school_id = school_data["organisationunitid_original"]
        dataset_uid = school_data.get("datasetuid", "")
        dataset_name = school_data.get("datasets_name", "")
        
        school_users = users_df[users_df["schoolUID"] == school_id].copy()
        school_users["schoolUID"] = school_users["schoolUID_original"]
        
        if school_users.empty:
            state = school_data.get("orgunitlevel2", "")
            lga = school_data.get("orgunitlevel3", "")
            ward = school_data.get("orgunitlevel4", "")
            status = "🚫 No user account"
            last_login = ""
        else:
            first_user = school_users.iloc[0]
            state = school_data.get("orgunitlevel2", "")
            lga = school_data.get("orgunitlevel3", "")
            ward = school_data.get("orgunitlevel4", "")
            has_login = school_users["lastLogin"].str.strip().ne("").any()
            has_upload = school_id in uploaded_schools
            if has_upload:
                status = "✅ Logged in and uploaded data"
            elif has_login:
                status = "⚠️ Logged in, no data upload"
            else:
                status = "❌ User exists, yet to login"
            if len(school_users) > 1:
                status += f" (👥 {len(school_users)} users)"
            last_logins = pd.to_datetime(school_users["lastLogin"], errors='coerce', utc=True).dropna()
            last_login = last_logins.max().strftime('%Y-%m-%d %H:%M:%S UTC') if not last_logins.empty else ""
        
        report_rows.append({
            "State": state,
            "LGA": lga,
            "Ward": ward,
            "School ID": original_school_id,
            "School Name": school_name,
            "datasetuid": dataset_uid,
            "datasets_name": dataset_name,
            "Status": status,
            "User Count": len(school_users),
            "Usernames": ", ".join(school_users["username"]),
            "Last Login": last_login
        })
    
    report_df = pd.DataFrame(report_rows)
    
    # Add Validation Rule Violations only for ⚠️ Logged in, no data upload with batching
    violations_df = pd.DataFrame(columns=["State", "LGA", "Ward", "School ID", "School Name", "Dataset UID", "datasets_name", "Period", "Validation Rule Name", "Left Side Value", "Right Side Value", "Importance"])
    no_upload_rows = report_df[report_df["Status"].str.startswith("⚠️ Logged in, no data upload")]
    if not no_upload_rows.empty:
        print(f"Processing {len(no_upload_rows)} rows with ⚠️ status")
        batch_size = int(request.args.get('batchSize', 10))  # Default to 10, options: 5, 10, 15
        
        # Deduplicate by School ID and collect all dataset_uids
        unique_ous = no_upload_rows.groupby("School ID")["datasetuid"].apply(list).reset_index()
        batches = [unique_ous.iloc[i:i + batch_size][["School ID", "datasetuid"]].values.tolist() for i in range(0, len(unique_ous), batch_size)]
        
        with ThreadPoolExecutor(max_workers=4) as executor:
            future_to_batch = {executor.submit(get_validation_violations_batch, batch, report_df): batch for batch in batches}
            for future in as_completed(future_to_batch):
                batch = future_to_batch[future]
                try:
                    results = future.result()
                    for (orgunit_id, dataset_uids), result in results.items():
                        violations = result["violations"]
                        state = result["state"]
                        lga = result["lga"]
                        ward = result["ward"]
                        dataset_name = result["dataset_name"]
                        orgunit_name = get_orgunit_name(orgunit_id)
                        if not violations:
                            print("   ℹ️ No violations found.")
                        else:
                            print(f"   ⚠️ {len(violations)} violation(s) found:")
                            for v in violations:
                                rule = v.get("validationRule", {})
                                rule_name = rule.get("name", "Unnamed Rule")
                                print(f"     - {rule_name}")
                                violations_df = pd.concat([violations_df, pd.DataFrame([{
                                    "State": state,
                                    "LGA": lga,
                                    "Ward": ward,
                                    "School ID": orgunit_id,
                                    "School Name": orgunit_name,
                                    "Dataset UID": v.get("dataSet", {}).get("id", dataset_uids[0]),  # Fallback to first dataset_uid if not in response
                                    "datasets_name": dataset_name,
                                    "Period": v.get("period", {}).get("name", "2024"),
                                    "Validation Rule Name": rule_name,
                                    "Left Side Value": v.get("leftsideValue"),
                                    "Right Side Value": v.get("rightsideValue"),
                                    "Importance": rule.get("importance", "N/A")
                                }])], ignore_index=True)
                except Exception as e:
                    print(f"⚠️ Batch processing failed: {e}")
    
    # Update Excel file with violations
    temp_excel = os.path.join(OUTPUT_FOLDER, "temp_violations.xlsx")
    try:
        if os.path.exists(REPORT_EXCEL):
            shutil.copy(REPORT_EXCEL, temp_excel)
            wb = load_workbook(temp_excel)
        else:
            # Create a new workbook if report doesn't exist
            report_df.to_excel(temp_excel, index=False, sheet_name="Report")
            wb = load_workbook(temp_excel)
            ws_report = wb["Report"]
            for cell in ws_report[1]:
                cell.font = Font(bold=True)
            for column_cells in ws_report.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                ws_report.column_dimensions[column_cells[0].column_letter].width = max_length + 2
            status_colors = {
                "✅ Logged in and uploaded data": "C6EFCE",
                "⚠️ Logged in, no data upload": "FFF2CC",
                "❌ User exists, yet to login": "F8CBAD",
                "🚫 No user account": "D9D9D9"
            }
            status_col = [cell.value for cell in ws_report[1]].index("Status") + 1
            for row in ws_report.iter_rows(min_row=2, max_row=ws_report.max_row):
                status = row[status_col - 1].value
                if isinstance(status, str):
                    for key, color in status_colors.items():
                        if status.startswith(key):
                            fill = PatternFill(start_color=color, fill_type="solid")
                            for cell in row:
                                cell.fill = fill
                            break
        
        if not violations_df.empty:
            if "Violations" in wb.sheetnames:
                ws_violations = wb["Violations"]
                ws_violations.delete_rows(2, ws_violations.max_row)  # Clear existing data
            else:
                ws_violations = wb.create_sheet("Violations")
                headers = ["State", "LGA", "Ward", "School ID", "School Name", "Dataset UID", "datasets_name", "Period", "Validation Rule Name", "Left Side Value", "Right Side Value", "Importance"]
                ws_violations.append(headers)
                for cell in ws_violations[1]:
                    cell.font = Font(bold=True)
            
            for _, row in violations_df.iterrows():
                ws_violations.append([row[col] for col in headers])
            
            for column_cells in ws_violations.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                ws_violations.column_dimensions[column_cells[0].column_letter].width = max_length + 2
        
        wb.save(temp_excel)
        if os.path.exists(REPORT_EXCEL):
            os.remove(REPORT_EXCEL)
        shutil.move(temp_excel, REPORT_EXCEL)
    except Exception as e:
        print(f"⚠️ Failed to update Excel file with violations: {e}")
        if os.path.exists(temp_excel):
            os.remove(temp_excel)
        raise

    return jsonify({"violations": violations_df.to_dict(orient='records')})

# Download files
@app.route('/api/download/<filename>', methods=['GET'])
def download_file(filename):
    file_map = {
        "report.xlsx": REPORT_EXCEL,
        "no_upload.xlsx": NO_UPLOAD_FILE,
        "chart.png": STATUS_CHART_PATH
    }
    if filename in file_map and os.path.exists(file_map[filename]):
        return send_file(file_map[filename], as_attachment=True)
    return jsonify({"error": "File not found"}), 404

# Serve frontend
@app.route('/')
def serve_index():
    return send_from_directory('.', 'index.html')

if __name__ == "__main__":
    app.run(debug=True, port=5000)